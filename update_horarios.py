"""Descarga el Excel de OneDrive, aplica los turnos del mensaje de WhatsApp
a la hoja 'Horarios Personal', y sube el Excel actualizado.

Variables de entorno necesarias:
  MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET  (Azure AD app con Files.ReadWrite.All)
  ONEDRIVE_SHARE_URL   (share URL del xlsx)
  MENSAJES_WHATSAPP    (texto crudo pegado por el user)
  [opcional] XLSX_SHEET  default 'Horarios Personal'
  [opcional] DRY_RUN=1   no sube cambios, solo imprime plan
"""
from __future__ import annotations

import base64
import copy
import datetime
import io
import os
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from whatsapp_parser import parse_mensajes, Turno

SHEET_NAME = os.environ.get("XLSX_SHEET", "Horarios Personal")
DRY_RUN    = os.environ.get("DRY_RUN", "").strip() == "1"

# Graph -----------------------------------------------------------------

def _graph_token() -> str:
    resp = requests.post(
        f"https://login.microsoftonline.com/{os.environ['MS_TENANT_ID']}/oauth2/v2.0/token",
        data={
            "client_id":     os.environ["MS_CLIENT_ID"],
            "client_secret": os.environ["MS_CLIENT_SECRET"],
            "scope":         "https://graph.microsoft.com/.default",
            "grant_type":    "client_credentials",
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def _encode_share_url(u: str) -> str:
    return "u!" + base64.urlsafe_b64encode(u.encode()).decode().rstrip("=")


def _share_item_url() -> str:
    enc = _encode_share_url(os.environ["ONEDRIVE_SHARE_URL"])
    return f"https://graph.microsoft.com/v1.0/shares/{enc}/driveItem"


def print_file_info(token: str) -> None:
    try:
        headers = {"Authorization": f"Bearer {token}"}
        r = requests.get(_share_item_url(), headers=headers, timeout=30)
        r.raise_for_status()
        it = r.json()
        name    = it.get("name", "?")
        size_kb = (it.get("size", 0) or 0) // 1024
        mod_at  = it.get("lastModifiedDateTime", "?")
        mod_by  = (it.get("lastModifiedBy", {}) or {}).get("user", {}).get("displayName", "?")
        parent  = (it.get("parentReference", {}) or {}).get("path", "?")
        print(f"\U0001f4c1 Fichero OneDrive: {name} ({size_kb} KB)")
        print(f"   Carpeta: {parent}")
        print(f"   Ultima edicion: {mod_at} por {mod_by}")
    except Exception as e:
        print(f"\u26a0 No se pudieron leer metadatos del driveItem: {e}")


def download_excel(token: str, dst: Path) -> None:
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(
        _share_item_url() + "/content",
        headers=headers, stream=True, allow_redirects=True, timeout=60,
    )
    r.raise_for_status()
    with open(dst, "wb") as f:
        for chunk in r.iter_content(65536):
            f.write(chunk)


def upload_excel(token: str, src: Path) -> None:
    """Resuelve el driveItem y hace PUT del contenido.
    Reintenta hasta 5 veces si OneDrive devuelve 423 (archivo bloqueado).
    """
    headers = {"Authorization": f"Bearer {token}"}
    meta = requests.get(_share_item_url(), headers=headers, timeout=30)
    meta.raise_for_status()
    item     = meta.json()
    drive_id = item["parentReference"]["driveId"]
    item_id  = item["id"]
    put_url  = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    with open(src, "rb") as f:
        data = f.read()
    esperas = [15, 30, 60, 120, 300]
    for intento, espera in enumerate(esperas, start=1):
        r = requests.put(
            put_url,
            headers={**headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            data=data,
            timeout=120,
        )
        if r.status_code == 423:
            print(f"\u26a0  OneDrive bloqueado (423) - intento {intento}/{len(esperas)}. Esperando {espera}s...")
            time.sleep(espera)
            continue
        r.raise_for_status()
        return
    r.raise_for_status()


# Excel -----------------------------------------------------------------

def _as_date(v) -> datetime.date | None:
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date):     return v
    return None


def _last_data_row(ws) -> int:
    last = 4
    for r in range(5, ws.max_row + 1):
        if ws.cell(r, 1).value is not None:
            last = r
    return last


def _existing_fecha_turno(ws) -> set[tuple[datetime.date, str]]:
    pairs: set[tuple[datetime.date, str]] = set()
    for r in range(5, ws.max_row + 1):
        d = _as_date(ws.cell(r, 1).value)
        t = ws.cell(r, 3).value
        if d is None or t is None: continue
        pairs.add((d, str(t).strip()))
    return pairs


def _formula_template(ws, row: int) -> dict[int, str]:
    tpl: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row, col).value
        if isinstance(v, str) and v.startswith("="):
            tpl[col] = v
    return tpl


def _copy_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell.font          = copy.copy(src_cell.font)
        dst_cell.fill          = copy.copy(src_cell.fill)
        dst_cell.border        = copy.copy(src_cell.border)
        dst_cell.alignment     = copy.copy(src_cell.alignment)
        dst_cell.protection    = copy.copy(src_cell.protection)
        dst_cell.number_format = src_cell.number_format


def apply_turnos(ws, turnos: list[Turno]) -> tuple[list[Turno], list[Turno]]:
    existing   = _existing_fecha_turno(ws)
    tpl_row    = 5
    tpl        = _formula_template(ws, tpl_row)
    applied: list[Turno] = []
    skipped: list[Turno] = []
    next_row   = _last_data_row(ws) + 1
    tpl_height = ws.row_dimensions[tpl_row].height

    for t in turnos:
        if t.key() in existing:
            skipped.append(t)
            continue
        for col in range(1, ws.max_column + 1):
            _copy_style(ws.cell(tpl_row, col), ws.cell(next_row, col))
        if tpl_height is not None:
            ws.row_dimensions[next_row].height = tpl_height
        ws.cell(next_row, 1).value = datetime.datetime.combine(t.fecha, datetime.time(0, 0))
        ws.cell(next_row, 3).value = t.turno
        ws.cell(next_row, 4).value = t.nombre
        ws.cell(next_row, 5).value = t.entrada
        ws.cell(next_row, 6).value = t.salida
        for col, formula in tpl.items():
            if col in (1, 3, 4, 5, 6): continue
            ws.cell(next_row, col).value = Translator(
                formula, origin=f"A{tpl_row}"
            ).translate_formula(f"A{next_row}")
        applied.append(t)
        next_row += 1

    return applied, skipped


# Entry point -----------------------------------------------------------

def main() -> int:
    texto = os.environ.get("MENSAJES_WHATSAPP", "").strip()
    if not texto:
        print("ERROR: falta MENSAJES_WHATSAPP", file=sys.stderr)
        return 2

    turnos = parse_mensajes(texto)
    print(f"\U0001f4e8 Turnos detectados en el mensaje: {len(turnos)}")
    for t in turnos:
        print(f"   {t.fecha} {t.turno:<7} {t.nombre:<12} {t.entrada}-{t.salida}")

    if not turnos:
        print("No hay turnos que aplicar. Revisa el formato.")
        return 0

    print("\n\U0001f510 Autenticando con Microsoft Graph...")
    token = _graph_token()
    print_file_info(token)

    local = Path("/tmp/horarios.xlsx")
    print("\u2b07\ufe0f  Descargando Excel desde OneDrive...")
    download_excel(token, local)
    print(f"   {local.stat().st_size // 1024} KB")

    print(f"\n\U0001f4ca Abriendo hoja '{SHEET_NAME}'...")
    wb = load_workbook(local)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: la hoja '{SHEET_NAME}' no existe. Hojas: {wb.sheetnames}", file=sys.stderr)
        return 3

    ws = wb[SHEET_NAME]
    print(f"   Filas actuales: {_last_data_row(ws)}")

    applied, skipped = apply_turnos(ws, turnos)

    print(f"\n\u2705 Aplicados: {len(applied)}")
    for t in applied:
        print(f"   + {t.fecha} {t.turno:<7} {t.nombre:<12} {t.entrada}-{t.salida}")

    if skipped:
        print(f"\n\u23ed Omitidos (ya existen para esa fecha+turno): {len(skipped)}")
        keys = {t.key() for t in skipped}
        for k in sorted(keys):
            print(f"   - {k[0]} {k[1]}")

    if not applied:
        print("\nNada que subir.")
        return 0

    if DRY_RUN:
        print("\n\U0001f9ea DRY_RUN=1 -> no subimos el Excel.")
        return 0

    wb.save(local)
    print(f"\n\u2b06\ufe0f  Subiendo Excel actualizado a OneDrive ({local.stat().st_size // 1024} KB)...")
    upload_excel(token, local)
    print("\U0001f389 OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
