"""
update_facturacion.py
─────────────────────────────────────────────────────────────────────────────
Lee todos los archivos "Cierres de cajas*.xlsx" de la carpeta OneDrive,
extrae los totales por fecha y turno (Desayuno→Mañana, Comida→Mediodía,
Cena→Noche), y actualiza la hoja 'Facturación 2026' del Excel principal
en aquellas fechas donde el total actual es 0 (sin datos todavía).

Variables de entorno necesarias (igual que update_horarios.py):
  MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET
  ONEDRIVE_SHARE_URL  (share URL del Excel principal)
  [opcional] DRY_RUN=1  imprime cambios sin subir
"""
from __future__ import annotations

import base64
import datetime
import io
import os
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

# ── Configuración ─────────────────────────────────────────────────────────────

SHEET_FAC  = "Facturación 2026"
DRY_RUN    = os.environ.get("DRY_RUN", "").strip() == "1"

# Mapeo turno cierre → columna openpyxl en Facturación 2026
# col 3 = Mañana €, col 4 = Mediodía €, col 5 = Noche €, col 6 = Total Real €
TURNO_COL = {
    "desayuno": 3,   # → Mañana
    "comida":   4,   # → Mediodía
    "cena":     5,   # → Noche
}
COL_TOTAL = 6        # Total Real € (=SUM de Mañana+Mediodía+Noche)

# ── Microsoft Graph ───────────────────────────────────────────────────────────

def _graph_token() -> str:
    resp = requests.post(
        f"https://login.microsoftonline.com/{os.environ['MS_TENANT_ID']}/oauth2/v2.0/token",
        data={
            "client_id":     os.environ["MS_CLIENT_ID"],
            "client_secret": os.environ["MS_CLIENT_SECRET"],
            "scope":         "https://graph.microsoft.com/.default",
            "grant_type":    "client_credentials",
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def _encode_share_url(u: str) -> str:
    return "u!" + base64.urlsafe_b64encode(u.encode()).decode().rstrip("=")


def _share_item_url() -> str:
    enc = _encode_share_url(os.environ["ONEDRIVE_SHARE_URL"])
    return f"https://graph.microsoft.com/v1.0/shares/{enc}/driveItem"


def _get_main_item(token: str) -> dict:
    r = requests.get(_share_item_url(), headers={"Authorization": f"Bearer {token}"}, timeout=30)
    r.raise_for_status()
    return r.json()


def list_folder_files(token: str) -> list[dict]:
    """Lista todos los archivos de la carpeta padre del Excel principal."""
    item    = _get_main_item(token)
    drive_id   = item["parentReference"]["driveId"]
    parent_id  = item["parentReference"]["id"]
    headers    = {"Authorization": f"Bearer {token}"}
    files: list[dict] = []
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_id}/children?$top=200"
    while url:
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()
        files.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return files, drive_id


def download_item(token: str, drive_id: str, item_id: str) -> bytes:
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": f"Bearer {token}"},
        allow_redirects=True, stream=True, timeout=60,
    )
    r.raise_for_status()
    return r.content


def upload_excel(token: str, drive_id: str, item_id: str, data: bytes) -> None:
    """Sube el Excel con reintentos en caso de 423 (bloqueado)."""
    put_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    for intento, espera in enumerate([15, 30, 60, 120, 300], start=1):
        r = requests.put(put_url, headers=headers, data=data, timeout=120)
        if r.status_code == 423:
            print(f"⚠  OneDrive bloqueado (423) — intento {intento}/5. Esperando {espera}s...")
            time.sleep(espera)
            continue
        r.raise_for_status()
        return
    r.raise_for_status()


# ── Parsear cierres ───────────────────────────────────────────────────────────

def parse_cierres(raw: bytes) -> dict[datetime.date, dict[int, float]]:
    """
    Devuelve {fecha: {col_openpyxl: total}} con los datos del cierre.
    Si hay varios registros para el mismo día/turno, se suman (por si hay 2 locales).
    """
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    result: dict[datetime.date, dict[int, float]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 16:
            continue
        fecha_raw = row[0]
        turno_raw = row[3]
        total_raw = row[15]

        # Parsear fecha
        if isinstance(fecha_raw, datetime.datetime):
            fecha = fecha_raw.date()
        elif isinstance(fecha_raw, str):
            try:
                fecha = datetime.datetime.strptime(fecha_raw.strip(), "%d/%m/%Y").date()
            except ValueError:
                continue
        else:
            continue

        if not turno_raw or not isinstance(total_raw, (int, float)) or total_raw <= 0:
            continue

        col = TURNO_COL.get(str(turno_raw).strip().lower())
        if col is None:
            continue

        if fecha not in result:
            result[fecha] = {}
        result[fecha][col] = round(result[fecha].get(col, 0.0) + float(total_raw), 2)

    return result


# ── Actualizar Excel ──────────────────────────────────────────────────────────

def update_excel(excel_raw: bytes, cierres: dict) -> tuple[bytes, int]:
    """
    Abre el Excel, detecta filas con total=0 que tienen datos en cierres,
    escribe los valores de turno + total directamente, y devuelve el Excel modificado.
    """
    # 1) Leer valores actuales (data_only) para detectar qué filas tienen total=0
    wb_ro = load_workbook(io.BytesIO(excel_raw), read_only=True, data_only=True)
    ws_ro = wb_ro[SHEET_FAC]
    pendientes: dict[datetime.date, int] = {}   # {fecha: row_num_openpyxl}

    for i, row in enumerate(ws_ro.iter_rows(min_row=4, values_only=True), start=4):
        fecha_val = row[0]
        if not isinstance(fecha_val, datetime.datetime):
            continue
        fecha = fecha_val.date()
        if fecha not in cierres:
            continue
        total_actual = row[5]   # col F = Total Real €
        if not isinstance(total_actual, (int, float)) or total_actual == 0:
            pendientes[fecha] = i

    wb_ro.close()

    if not pendientes:
        return excel_raw, 0

    # 2) Abrir para escritura y rellenar las celdas
    wb = load_workbook(io.BytesIO(excel_raw))
    ws = wb[SHEET_FAC]
    updated = 0

    for fecha, row_num in sorted(pendientes.items()):
        day = cierres[fecha]
        ma  = day.get(3, 0.0)
        md  = day.get(4, 0.0)
        no  = day.get(5, 0.0)
        tot = round(ma + md + no, 2)

        ws.cell(row_num, 3).value = ma    # Mañana
        ws.cell(row_num, 4).value = md    # Mediodía
        ws.cell(row_num, 5).value = no    # Noche
        ws.cell(row_num, 6).value = tot   # Total Real (sobreescribimos =SUM)

        print(f"   ✅ {fecha.strftime('%d/%m/%Y')}  "
              f"MA={ma:>8.2f}  MD={md:>8.2f}  NO={no:>8.2f}  TOT={tot:>9.2f}")
        updated += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), updated


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> int:
    print("🔐 Autenticando con Microsoft Graph...")
    token = _graph_token()

    print("📂 Listando carpeta OneDrive...")
    files, drive_id = list_folder_files(token)
    cierres_files = [
        f for f in files
        if "cierres de cajas" in f["name"].lower() and f["name"].lower().endswith(".xlsx")
    ]
    print(f"   Archivos de cierres encontrados: {len(cierres_files)}")
    if not cierres_files:
        print("⚠  No hay archivos de cierres. Nada que hacer.")
        return 0

    # Parsear todos los cierres y consolidar
    all_cierres: dict[datetime.date, dict[int, float]] = {}
    for cf in cierres_files:
        print(f"   📥 Leyendo: {cf['name']}")
        raw = download_item(token, drive_id, cf["id"])
        day_data = parse_cierres(raw)
        for fecha, turnos in day_data.items():
            if fecha not in all_cierres:
                all_cierres[fecha] = {}
            for col, total in turnos.items():
                # Si hay solapamiento entre archivos, conserva el valor mayor
                all_cierres[fecha][col] = max(all_cierres[fecha].get(col, 0.0), total)

    print(f"   Total días con datos de cierre: {len(all_cierres)}")

    print("\n⬇  Descargando Excel principal...")
    item       = _get_main_item(token)
    item_id    = item["id"]
    excel_raw  = download_item(token, drive_id, item_id)
    print(f"   {len(excel_raw) // 1024} KB")

    print(f"\n📊 Actualizando hoja '{SHEET_FAC}'...")
    new_excel, n_updated = update_excel(excel_raw, all_cierres)

    if n_updated == 0:
        print("✅ Todas las fechas ya tienen datos. Nada que actualizar.")
        return 0

    print(f"\n   {n_updated} día(s) actualizado(s).")

    if DRY_RUN:
        print("🧪 DRY_RUN=1 — no se sube el Excel.")
        return 0

    print(f"\n⬆  Subiendo Excel actualizado ({len(new_excel) // 1024} KB)...")
    upload_excel(token, drive_id, item_id, new_excel)
    print("🎉 OK — facturación actualizada en OneDrive.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
