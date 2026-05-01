#!/usr/bin/env python3
"""
Taperia de Caldes — Generador Tracker Operativo 2026
Lee el Excel, extrae todos los datos y los inyecta en tracker-template.html
como 'const DATA = {...}' → genera tracker-operativo.html

El diseño/CSS/estructura HTML permanece FIJO en tracker-template.html.
Solo los datos cambian en cada ejecución.

Uso: python3 generar_tracker.py [ruta_excel]
"""
import sys
import os
import json
import re
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: Instala openpyxl con: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None

# ── Descarga automática desde OneDrive (para GitHub Actions) ─────────────────
def _download_excel_onedrive(dest_path: str) -> bool:
    """Descarga el Excel desde OneDrive si están definidas las variables de entorno."""
    tenant    = os.environ.get("MS_TENANT_ID", "").strip()
    client_id = os.environ.get("MS_CLIENT_ID", "").strip()
    secret    = os.environ.get("MS_CLIENT_SECRET", "").strip()
    share_url = os.environ.get("ONEDRIVE_SHARE_URL", "").strip()
    if not (tenant and client_id and secret and share_url):
        return False
    if requests is None:
        print("ERROR: pip install requests  (necesario para descarga OneDrive)")
        return False
    import base64
    print("🔐 Autenticando con Microsoft Graph para descarga del Excel...")
    r = requests.post(
        f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials", "client_id": client_id,
              "client_secret": secret, "scope": "https://graph.microsoft.com/.default"},
        timeout=30
    )
    r.raise_for_status()
    token = r.json()["access_token"]
    enc = "u!" + base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    print("⬇  Descargando Excel desde OneDrive...")
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{enc}/driveItem/content",
        headers={"Authorization": f"Bearer {token}"},
        stream=True, allow_redirects=True, timeout=60
    )
    r.raise_for_status()
    with open(dest_path, "wb") as f:
        for chunk in r.iter_content(65536):
            f.write(chunk)
    print(f"   ✅ Excel descargado: {os.path.getsize(dest_path)//1024} KB → {dest_path}")
    return True

# ── Localizar archivos ────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

if len(sys.argv) > 1:
    EXCEL_PATH = sys.argv[1]
else:
    # En GitHub Actions: intentar descarga automática si hay env vars de OneDrive
    auto_dest = os.path.join(SCRIPT_DIR, "Taperia_Caldes_Operaciones_2026_Trabajando.xlsx")
    if os.environ.get("MS_TENANT_ID") and not os.path.exists(auto_dest):
        if not _download_excel_onedrive(auto_dest):
            print("ERROR: No se pudo descargar el Excel desde OneDrive.")
            sys.exit(1)

    candidates = [f for f in os.listdir(SCRIPT_DIR) if 'Taperia' in f and f.endswith('.xlsx')]
    if not candidates:
        print("ERROR: No se encontró el Excel. Pásalo como argumento:")
        print("  python3 generar_tracker.py Taperia_Caldes_Operaciones_2026.xlsx")
        sys.exit(1)
    # Preferir el archivo "Trabajando" (el activo) sobre los OLD
    trabajando = [f for f in candidates if 'Trabajando' in f]
    EXCEL_PATH = os.path.join(SCRIPT_DIR, trabajando[0] if trabajando else sorted(candidates)[-1])

TEMPLATE = os.path.join(SCRIPT_DIR, 'tracker-template.html')
OUTPUT   = os.path.join(SCRIPT_DIR, 'tracker-operativo.html')

print(f"Leyendo: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ── Helpers ──────────────────────────────────────────────────────────────────
mes_names = {
    1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio',
    7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'
}

# ── Facturación 2026 ──────────────────────────────────────────────────────────
ws_fac = wb['Facturación 2026']
days = []
for row in ws_fac.iter_rows(min_row=4, values_only=True):
    fecha = row[0]
    if not isinstance(fecha, datetime): continue
    real = row[5]
    if not isinstance(real, (int, float)) or real <= 0: continue
    coste_p = row[9] if isinstance(row[9], (int, float)) else 0
    days.append({
        'fecha':    fecha.strftime('%d/%m'),
        'dia':      (row[1] or '').capitalize(),
        'mes':      fecha.month,
        'sem':      fecha.isocalendar()[1],
        'real':     round(float(real), 0),
        'previsto': round(float(row[6] or 0), 0),
        'desv_pct': round(float(row[8] or 0)*100, 1) if isinstance(row[8], (int, float)) else 0,
        'coste_p':  round(float(coste_p), 0),
        'evento':   str(row[11]) if row[11] else '',
        'manana':   round(float(row[2] or 0), 0),
        'mediodia': round(float(row[3] or 0), 0),
        'noche':    round(float(row[4] or 0), 0),
    })

mes_totals = {}
for d in days:
    m = d['mes']
    if m not in mes_totals:
        mes_totals[m] = {'real': 0, 'previsto': 0, 'coste': 0, 'dias': 0}
    mes_totals[m]['real']     += d['real']
    mes_totals[m]['previsto'] += d['previsto']
    mes_totals[m]['coste']    += d['coste_p']
    mes_totals[m]['dias']     += 1

total_real = sum(d['real'] for d in days)
total_prev = sum(d['previsto'] for d in days)
total_dias = len(days)

dow_totals = defaultdict(lambda: {'real': 0, 'dias': 0, 'coste': 0})
for d in days:
    dow_totals[d['dia']]['real']  += d['real']
    dow_totals[d['dia']]['dias']  += 1
    dow_totals[d['dia']]['coste'] += d['coste_p']

top10     = sorted(days, key=lambda x: -x['real'])[:10]
ultimos15 = days[-15:][::-1]

coste_real  = sum(v['coste'] for v in mes_totals.values())
pct_personal = round(coste_real / total_real * 100, 1) if total_real else 0
margen       = round(total_real - coste_real, 0)
pct_vs_prev  = round((total_real / total_prev - 1) * 100, 1) if total_prev else 0

# Ticket medio por turno (media de facturación en días con ese turno activo)
dias_m  = [d for d in days if d['manana']   > 0]
dias_md = [d for d in days if d['mediodia'] > 0]
dias_n  = [d for d in days if d['noche']    > 0]
ticket_medio = {
    'global':   round(total_real / total_dias, 0) if total_dias else 0,
    'manana':   round(sum(d['manana']   for d in dias_m)  / len(dias_m),  0) if dias_m  else 0,
    'mediodia': round(sum(d['mediodia'] for d in dias_md) / len(dias_md), 0) if dias_md else 0,
    'noche':    round(sum(d['noche']    for d in dias_n)  / len(dias_n),  0) if dias_n  else 0,
    'dias_m':   len(dias_m),
    'dias_md':  len(dias_md),
    'dias_n':   len(dias_n),
}

# Rendimiento por día de semana
dow_order  = ['Sáb', 'Dom', 'Vie', 'Jue', 'Mié']
best_media = max(
    (float(dow_totals[d]['real']) / max(dow_totals[d]['dias'], 1)
     for d in dow_order if dow_totals[d]['dias'] > 0),
    default=1
)
dow_list = []
for dia in dow_order:
    dv = dow_totals[dia]
    if dv['dias'] == 0: continue
    media = float(dv['real']) / dv['dias']
    pc    = float(dv['coste']) / float(dv['real']) * 100 if dv['real'] else 0
    bcls  = 'bar-g' if pc < 32 else ('bar-o' if pc < 40 else 'bar-r')
    dow_list.append({
        'dia': dia, 'dias': dv['dias'],
        'total':        round(float(dv['real']), 0),
        'media':        round(media, 0),
        'pct_personal': round(pc, 1),
        'bar_cls':      bcls,
        'bar_w':        round(min(100, media / best_media * 100), 1) if best_media else 0,
    })

# Datos para los bloques de datos (top10, ultimos15)
def day_to_dict(d):
    return {k: d[k] for k in ('fecha', 'dia', 'real', 'previsto', 'desv_pct', 'evento')}

# ── Horarios Personal ─────────────────────────────────────────────────────────
ws_hor = wb['Horarios Personal']
empleados    = defaultdict(lambda: {'coste': 0, 'horas': 0})
semanas_data = defaultdict(lambda: {'coste': 0, 'ventas': 0})
meses_personal = defaultdict(float)

for d in days:
    semanas_data[d['sem']]['ventas'] += d['real']

for row in ws_hor.iter_rows(min_row=5, values_only=True):
    fecha   = row[0]
    if not isinstance(fecha, datetime): continue
    nombre  = row[3]
    coste_t = row[11]
    horas   = row[6]
    if not nombre or not isinstance(coste_t, (int, float)): continue
    empleados[nombre]['coste'] += coste_t
    empleados[nombre]['horas'] += (horas or 0) if isinstance(horas, (int, float)) else 0
    semanas_data[fecha.isocalendar()[1]]['coste'] += coste_t
    meses_personal[fecha.month] += coste_t

total_coste_emp = sum(v['coste'] for v in empleados.values())
emp_sorted = sorted(empleados.items(), key=lambda x: -x[1]['coste'])
mx_emp = float(emp_sorted[0][1]['coste']) if emp_sorted else 1

emp_list = []
for n, v in emp_sorted:
    c = float(v['coste']); h = float(v['horas'])
    emp_list.append({
        'nombre':     n,
        'coste':      round(c, 0),
        'pct_masa':   round(c / total_coste_emp * 100, 1) if total_coste_emp else 0,
        'horas':      round(h, 0),
        'coste_hora': round(c / h, 2) if h else None,
        'bar_w':      round(min(100, c / mx_emp * 100), 1) if mx_emp else 0,
    })

sems_list = []
cum_c = 0; cum_v = 0
for s in sorted(semanas_data.keys()):
    c   = semanas_data[s]['coste']
    v_s = semanas_data[s]['ventas']
    cum_c += c; cum_v += v_s
    pct_s   = round(c / v_s * 100, 1) if v_s else 0
    cum_pct = round(cum_c / cum_v * 100, 1) if cum_v else 0
    bcls = 'bar-g' if pct_s < 32 else ('bar-o' if pct_s < 40 else 'bar-r')
    sems_list.append({
        'sem':     s,
        'coste':   round(c, 0),
        'ventas':  round(v_s, 0),
        'pct':     pct_s,
        'cum_pct': cum_pct,
        'bar_cls': bcls,
        'bar_w':   round(min(pct_s, 70) / 70 * 100, 1) if pct_s else 0,
    })

meses_per_list = []
for m in range(1, 5):
    c  = float(meses_personal.get(m, 0))
    vv = float(mes_totals.get(m, {'real': 0})['real'])
    meses_per_list.append({
        'nombre': mes_names[m],
        'coste':  round(c, 0),
        'ventas': round(vv, 0),
        'pct':    round(c / vv * 100, 1) if vv else 0,
    })

# ── Eventos y Previsión ───────────────────────────────────────────────────────
ws_ev = wb['Eventos y Previsión']
eventos = []
for row in ws_ev.iter_rows(min_row=4, max_row=20, values_only=True):
    mes_ev, nombre_ev, tipo_ev = row[0], row[2], row[3]
    mult, prev_ev, real_ev, desv_ev, desv_pct_ev, estado = row[4], row[5], row[6], row[7], row[8], row[9]
    if not nombre_ev or not mes_ev: continue
    if isinstance(prev_ev, str) or isinstance(real_ev, str): continue
    estado_s = str(estado or '')
    est_cls  = ('bg' if 'Pasado' in estado_s else
                'bb' if 'Confirmado' in estado_s else
                'bo' if 'confirmar' in estado_s.lower() else 'bgr')
    estado_c = estado_s.replace('✓','').replace('⚡','').replace('📅','').replace('⚠️','').strip()
    eventos.append({
        'mes':      mes_ev,
        'nombre':   nombre_ev,
        'tipo':     tipo_ev or '',
        'mult':     mult or '',
        'prev':     round(float(prev_ev), 0) if isinstance(prev_ev, (int, float)) else None,
        'real':     round(float(real_ev), 0) if isinstance(real_ev, (int, float)) else None,
        'desv':     round(float(desv_ev), 0) if isinstance(desv_ev, (int, float)) else None,
        'desv_pct': round(float(desv_pct_ev) * 100, 1) if isinstance(desv_pct_ev, (int, float)) else None,
        'estado':   estado_c,
        'est_cls':  est_cls,
    })

# ── Dashboard Gerencia ────────────────────────────────────────────────────────
ws_dash = wb['Dashboard Gerencia']
alertas  = []
kpis_fin = []
dow_efic = []

for row in ws_dash.iter_rows(min_row=47, max_row=56, values_only=True):
    prio, concepto, detalle = row[1], row[3], row[6]
    if not prio or prio == 'Prioridad': continue
    alertas.append({
        'prio':     str(prio),
        'concepto': str(concepto or '').strip(),
        'detalle':  str(detalle or '').strip(),
    })

for row in ws_dash.iter_rows(min_row=58, max_row=67, values_only=True):
    kpi, val, obj, estado, tend = row[0], row[1], row[2], row[3], row[4]
    if not kpi or kpi == 'KPI Financiero': continue
    val_fmt = (f"{val*100:.1f}%" if isinstance(val, float) and val < 10 else
               f"{val:,.0f}€".replace(',', '.') if isinstance(val, (int, float)) else str(val or ''))
    est_cls = 'bg' if 'OK' in str(estado) else ('bo' if 'Atención' in str(estado) else 'br')
    kpis_fin.append({
        'kpi':      kpi,
        'val':      val_fmt,
        'obj':      str(obj or ''),
        'estado':   str(estado or '').replace('🟢','').replace('🟡','').replace('🔴','').strip(),
        'est_cls':  est_cls,
        'tend':     str(tend or ''),
    })

for row in ws_dash.iter_rows(min_row=27, max_row=33, values_only=True):
    dia, fact, coste, ratio, obj, estado, rec = row[0], row[1], row[2], row[3], row[5], row[6], row[8]
    if not dia or not isinstance(fact, (int, float)) or fact == 0: continue
    est_cls = 'bg' if 'OK' in str(estado) else 'bo'
    dow_efic.append({
        'dia':    dia,
        'fact':   round(float(fact), 0),
        'coste':  round(float(coste), 0),
        'ratio':  round(float(ratio) * 100, 1) if isinstance(ratio, float) else 0,
        'obj':    round(float(obj or 0) * 100, 0),
        'estado': str(estado or '').replace('🟢','').replace('🟡','').replace('🔴','').strip(),
        'est_cls':est_cls,
    })

# ── Product Mix Unificado ─────────────────────────────────────────────────────
ws_pm    = wb['Product Mix Unificado']
productos = defaultdict(int)
familias  = defaultdict(int)
for row in ws_pm.iter_rows(min_row=2, values_only=True):
    if row[2] and isinstance(row[4], (int, float)):
        productos[row[2]] += int(row[4])
        if row[3]: familias[row[3]] += int(row[4])

top_prods = sorted(productos.items(), key=lambda x: -x[1])[:20]
top_fams  = sorted(familias.items(),  key=lambda x: -x[1])[:12]
max_prod  = top_prods[0][1] if top_prods else 1
max_fam   = top_fams[0][1]  if top_fams  else 1
tot_fam_u = sum(u for _, u in top_fams)

prod_list = [
    {'rank': i+1, 'nombre': p, 'uds': u,
     'bar_w': round(min(100, u / max_prod * 100), 1)}
    for i, (p, u) in enumerate(top_prods)
]
fam_list = [
    {'rank': i+1, 'nombre': f, 'uds': u,
     'pct':   round(u / tot_fam_u * 100, 1) if tot_fam_u else 0,
     'bar_w': round(min(100, u / max_fam * 100), 1)}
    for i, (f, u) in enumerate(top_fams)
]

# ── Análisis Omnes ────────────────────────────────────────────────────────────
ws_om = wb['Análisis Omnes']
omnes = []
for row in ws_om.iter_rows(min_row=6, max_row=24, values_only=True):
    sf = row[0]
    if not sf: continue
    ap_cls = 'bg' if 'OK' in str(row[9])  else 'br'
    do_cls = ('bg' if 'OK'   in str(row[11]) else
              'bo' if 'Alto' in str(row[11]) else 'br')
    omnes.append({
        'sf':       sf,
        'np':       row[1],
        'pmin':     row[2],
        'pmax':     row[3],
        'baja':     row[4],
        'media_p':  row[5],
        'alta':     row[6],
        'aper':     round(float(row[7]), 2) if isinstance(row[7], (int, float)) else row[7],
        'lim':      row[8],
        'est_ap':   str(row[9]  or '').replace('✅','').replace('❌','').strip(),
        'ap_cls':   ap_cls,
        'ratio_do': row[10],
        'est_do':   str(row[11] or '').replace('✅','').replace('❌','').strip(),
        'do_cls':   do_cls,
    })

# ── P&L EBITDA ────────────────────────────────────────────────────────────────
ws_pl = wb['P&L EBITDA']
pl_excel = {}
for row in ws_pl.iter_rows(min_row=4, max_row=60, values_only=True):
    concepto = row[1]
    if not concepto or '%' in str(concepto): continue
    pl_excel[concepto] = {
        'ene':  float(row[4] or 0),
        'feb':  float(row[5] or 0),
        'mar':  float(row[6] or 0),
        'acum': float(row[2] or 0),
    }

# Parchar desde datos de facturación si P&L tiene 0
for m_num, m_key in [(1,'ene'), (2,'feb'), (3,'mar')]:
    mt = mes_totals.get(m_num, {})
    if pl_excel.get('Ventas', {}).get(m_key, 0) == 0 and mt.get('real', 0):
        pl_excel.setdefault('Ventas', {})[m_key] = mt['real']
    if pl_excel.get('Personal', {}).get(m_key, 0) == 0 and meses_personal.get(m_num, 0):
        pl_excel.setdefault('Personal', {})[m_key] = meses_personal[m_num]

def pv(c, m): return float(pl_excel.get(c, {}).get(m, 0) or 0)
def pp(c, m):
    ref = pv('Ventas', m)
    return round(pv(c, m) / ref * 100, 1) if ref else 0.0

v_mar = pv('Ventas','mar'); v_feb = pv('Ventas','feb')
v_ene = pv('Ventas','ene'); v_acum_pl = pv('Ventas','acum')
# Si acum es 0 (fórmulas no cacheadas), usar suma de meses
if v_acum_pl == 0: v_acum_pl = v_mar + v_feb + v_ene

e_mar  = pv('S.O.P /EBITDA','mar');  e_feb  = pv('S.O.P /EBITDA','feb')
e_ene  = pv('S.O.P /EBITDA','ene');  e_acum = pv('S.O.P /EBITDA','acum')
if e_acum == 0 and any([e_mar, e_feb, e_ene]):
    e_acum = e_mar + e_feb + e_ene

def pp_acum(c):
    return round(pv(c,'acum') / v_acum_pl * 100, 1) if v_acum_pl else 0.0

# Construir filas del P&L
pl_rows = []
def pl_section(label):
    pl_rows.append({'section': True, 'label': label})
def pl_row(c, label=None, bold=False, subtotal=False):
    pl_rows.append({
        'label':    label or c,
        'bold':     bold,
        'subtotal': subtotal,
        'mar':   {'v': round(pv(c,'mar'),0),  'p': pp(c,'mar')},
        'feb':   {'v': round(pv(c,'feb'),0),  'p': pp(c,'feb')},
        'ene':   {'v': round(pv(c,'ene'),0),  'p': pp(c,'ene')},
        'acum':  {'v': round(pv(c,'acum'),0), 'p': pp_acum(c)},
    })

pl_section("INGRESOS")
pl_row('Ventas', bold=True)
pl_section("COSTES VARIABLES")
pl_row('CMV', 'CMV (coste mercancia vendida)')
pl_row('Producto Limpieza', 'Producto limpieza')
pl_row('Servicio en Mesa', 'Servicio en mesa')
pl_row('Otros', 'Otros variables')
pl_section("GASTOS ESTRUCTURALES")
pl_row('Personal')
pl_row('Suministros')
pl_row('Contratos')
pl_row('Marketing')
pl_row('Otros gastos')
pl_row('Mantenimiento')
pl_row('S.O.P.1 /EBITDAR', 'S.O.P. / EBITDAR', subtotal=True)
pl_section("COSTES FIJOS")
pl_row('Alquileres')
pl_row('Royalty')
pl_row('S.O.P /EBITDA', 'EBITDA', bold=True)

# ── Armar objeto DATA ─────────────────────────────────────────────────────────
now = datetime.now().strftime('%d/%m/%Y %H:%M')

meses_list = []
for m in range(1, 5):
    mt = mes_totals.get(m, {'real': 0, 'previsto': 0, 'coste': 0, 'dias': 0})
    r  = float(mt['real']); p = float(mt['previsto'])
    c  = float(mt['coste']); di = int(mt['dias'])
    dv = round((r / p - 1) * 100, 1) if p else 0
    meses_list.append({
        'nombre': mes_names[m], 'real': round(r, 0), 'previsto': round(p, 0),
        'dias': di, 'coste_p': round(c, 0), 'dv': dv,
    })

DATA = {
    "meta": {
        "generado": now,
        "periodo": "Ene-Abr",
    },
    "fac": {
        "total_real":   total_real,
        "total_prev":   total_prev,
        "total_dias":   total_dias,
        "pct_vs_prev":  pct_vs_prev,
        "pct_personal": pct_personal,
        "margen":       margen,
        "ticket_medio": ticket_medio,
        "meses":        meses_list,
        "top10":        [day_to_dict(d) for d in top10],
        "ultimos15":    [day_to_dict(d) for d in ultimos15],
        "dow":          dow_list,
    },
    "per": {
        "total_coste": round(total_coste_emp, 0),
        "n_empleados": len(emp_sorted),
        "ratio":       round(total_coste_emp / total_real * 100, 1) if total_real else 0,
        "coste_dia":   round(total_coste_emp / total_dias, 0) if total_dias else 0,
        "semanas":     sems_list,
        "empleados":   emp_list,
        "meses":       meses_per_list,
        "dow_efic":    dow_efic,
    },
    "ev":  {"eventos": eventos},
    "al":  {"alertas": alertas, "kpis": kpis_fin},
    "pm":  {"top_prods": prod_list, "top_fams": fam_list},
    "om":  {"omnes": omnes},
    "pl":  {
        "e_mar":  round(e_mar, 0),  "e_feb":  round(e_feb, 0),
        "e_ene":  round(e_ene, 0),  "e_acum": round(e_acum, 0),
        "v_mar":  round(v_mar, 0),  "v_feb":  round(v_feb, 0),
        "v_ene":  round(v_ene, 0),  "v_acum": round(v_acum_pl, 0),
        "cmv_mar_pct": pp('CMV','mar'),
        "per_mar_pct": pp('Personal','mar'),
        "ebitda_mar_pct": pp('S.O.P /EBITDA','mar'),
        "ebitda_acum_pct": pp_acum('S.O.P /EBITDA'),
        "rows": pl_rows,
    },
}

# ── Inyectar en plantilla ─────────────────────────────────────────────────────
if not os.path.exists(TEMPLATE):
    print(f"ERROR: Plantilla no encontrada: {TEMPLATE}")
    print("Asegúrate de que tracker-template.html está en la misma carpeta.")
    sys.exit(1)

with open(TEMPLATE, 'r', encoding='utf-8') as f:
    tmpl = f.read()

data_json = json.dumps(DATA, ensure_ascii=False, separators=(',', ':'))

new_html = re.sub(
    r'const DATA\s*=\s*__TRACKER_DATA__\s*;',
    f'const DATA = {data_json};',
    tmpl
)

if new_html == tmpl:
    print("⚠  ATENCIÓN: No se encontró el marcador 'const DATA = __TRACKER_DATA__;' en la plantilla.")
    print("   Revisa tracker-template.html.")

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(new_html)

print(f"✅ Generado: {OUTPUT}  ({len(new_html) // 1024} KB)")
